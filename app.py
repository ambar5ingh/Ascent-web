"""
ASCENT — WRI India GHG Emissions & Scenario Planning Tool
IPCC 2019 / GPC Framework | AR6 GWP values
"""

from flask import (Flask, jsonify, request, render_template, send_file,
                   session, redirect, url_for, abort)
from functools import wraps
import io, json, math, os, re, threading
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import plotly
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from werkzeug.security import generate_password_hash, check_password_hash

# ─────────────────────────────────────────────────────────────────────────────
# All calculation constants — emission factors, GWP, climate-zone energy demand,
# decarbonisation cost tables, fuel conversion factors, the India city master,
# and so on — live in the `data` package. Editing values? Touch data/* not here.
# ─────────────────────────────────────────────────────────────────────────────
from data import (
    # GWP (AR6)
    GWP_CH4, GWP_N2O,
    # Sector emission factors (legacy flat dict + structured map)
    EF, EF_BY_SECTOR, ETHANOL_BLENDS,
    # Building, waste, IPPU, AFOLU defaults
    ENERGY_DEMAND, WW_MCF, SW_DOC, IPPU_EF,
    AFOLU_ENTERIC, AFOLU_MANURE_CH4,
    # Cost / strategy tables
    ABATEMENT_COST,
    # Fuel conversion table + helpers (was conv_factors.py)
    CONV_FACTORS, GENERIC_FUEL_EF, tj_factor, emit_for_fuel,
    # Legacy fuel→TJ shortcuts (kl_to_tj / t_to_tj / kg_to_tj)
    FUEL_CONV,
    # India city master (4,900+ urban local bodies)
    INDIA_CITIES,
)

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "ascent-wri-secret-2024")

# ═══════════════════════════════════════════════════════════════════════════════
# AUTH + PER-USER DATA STORE
# ═══════════════════════════════════════════════════════════════════════════════
#
# Lightweight file-backed store. No external DB dependency — keeps deployment
# identical to the current Azure Web App setup. Two JSON files:
#   data/users.json          → { "alice": "<pbkdf2-hash>" }
#   data/forms/<user>.json   → { "form": { ... questionnaire fields ... } }
#
# Username + password are allowed to be identical (per product spec).

import tempfile

# ─── Storage backend: S3 (persistent, survives redeploys) or local files ──────
#
# When ASCENT_S3_BUCKET is set (the App Runner / container deployment), user
# accounts and saved questionnaires live in S3 so they survive redeploys and
# instance replacement. Without it, we fall back to a writable local directory
# so the app still runs on a laptop or any plain host.
#
# The public functions (_load_users / _save_users / _load_user_form /
# _save_user_form / _delete_user_form) keep the same signatures either way, so
# nothing else in the app changes.

_S3_BUCKET = os.environ.get("ASCENT_S3_BUCKET", "").strip()
_S3_PREFIX = os.environ.get("ASCENT_S3_PREFIX", "ascent").strip().strip("/")
_STORE_LOCK = threading.Lock()

_s3 = None
if _S3_BUCKET:
    try:
        import boto3
        _s3 = boto3.client("s3")
    except Exception as e:
        # If boto3 is missing or credentials fail, log and fall back to local.
        print(f"[ascent] S3 requested but unavailable ({e}); using local storage")
        _s3 = None


def _init_store_dir():
    """Pick a writable local directory (used when S3 isn't configured, and for
    scratch even when it is)."""
    candidates = [
        os.environ.get("ASCENT_DATA_DIR"),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "_store"),
        os.path.join(tempfile.gettempdir(), "ascent_store"),
    ]
    for base in candidates:
        if not base:
            continue
        try:
            os.makedirs(os.path.join(base, "forms"), exist_ok=True)
            testfile = os.path.join(base, ".write_test")
            with open(testfile, "w") as f:
                f.write("ok")
            os.remove(testfile)
            return base
        except Exception:
            continue
    base = os.path.join(tempfile.gettempdir(), "ascent_store")
    os.makedirs(os.path.join(base, "forms"), exist_ok=True)
    return base


_STORE_DIR  = _init_store_dir()
_USERS_FILE = os.path.join(_STORE_DIR, "users.json")
_FORMS_DIR  = os.path.join(_STORE_DIR, "forms")

_USERNAME_RE = re.compile(r"^[A-Za-z0-9_.\-]{3,32}$")


# ── S3 key helpers ────────────────────────────────────────────────────────────
def _s3_key(*parts):
    return "/".join([_S3_PREFIX, *parts]) if _S3_PREFIX else "/".join(parts)

def _s3_get_json(key):
    try:
        obj = _s3.get_object(Bucket=_S3_BUCKET, Key=key)
        return json.loads(obj["Body"].read().decode("utf-8"))
    except _s3.exceptions.NoSuchKey:
        return None
    except Exception as e:
        # ClientError for a missing key on some setups
        if getattr(e, "response", {}).get("Error", {}).get("Code") in ("NoSuchKey", "404"):
            return None
        raise

def _s3_put_json(key, data):
    _s3.put_object(Bucket=_S3_BUCKET, Key=key,
                   Body=json.dumps(data, indent=2).encode("utf-8"),
                   ContentType="application/json")

def _s3_delete(key):
    try:
        _s3.delete_object(Bucket=_S3_BUCKET, Key=key)
    except Exception:
        pass


# ── Users ────────────────────────────────────────────────────────────────────
def _load_users():
    if _s3:
        return _s3_get_json(_s3_key("users.json")) or {}
    try:
        with open(_USERS_FILE, "r") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}

def _save_users(users):
    if _s3:
        _s3_put_json(_s3_key("users.json"), users)
        return
    tmp = _USERS_FILE + ".tmp"
    with open(tmp, "w") as f:
        json.dump(users, f, indent=2)
    os.replace(tmp, _USERS_FILE)


# ── Per-user saved questionnaire ──────────────────────────────────────────────
def _user_form_path(username):
    # username already validated against _USERNAME_RE, so it's safe
    return os.path.join(_FORMS_DIR, f"{username}.json")

def _load_user_form(username):
    if _s3:
        return _s3_get_json(_s3_key("forms", f"{username}.json")) or {}
    try:
        with open(_user_form_path(username), "r") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}

def _save_user_form(username, payload):
    if _s3:
        _s3_put_json(_s3_key("forms", f"{username}.json"), payload)
        return
    p = _user_form_path(username)
    tmp = p + ".tmp"
    with open(tmp, "w") as f:
        json.dump(payload, f, indent=2)
    os.replace(tmp, p)

def _delete_user_form(username):
    if _s3:
        _s3_delete(_s3_key("forms", f"{username}.json"))
        return
    try:
        os.remove(_user_form_path(username))
    except FileNotFoundError:
        pass

def current_user():
    return session.get("username")

def login_required(view_fn):
    """Gate a route behind authentication.
       Browser routes redirect to /login; API routes return 401 JSON."""
    @wraps(view_fn)
    def wrapper(*args, **kwargs):
        if not current_user():
            if request.path.startswith("/api/"):
                return jsonify({"error": "Not signed in"}), 401
            return redirect(url_for("login_page", next=request.path))
        return view_fn(*args, **kwargs)
    return wrapper

# Make the current username available to every template automatically,
# so the login pill + Logout link render in the header.
@app.context_processor
def _inject_current_user():
    return {"current_user": current_user()}


# ═══════════════════════════════════════════════════════════════════════════════
# CALCULATION ENGINE — mirrors Excel sheet-by-sheet logic
# ═══════════════════════════════════════════════════════════════════════════════

def ef_total(key):
    """Return total CO2e EF for a named fuel key."""
    e = EF[key]
    return e["co2"] + e["ch4"] * GWP_CH4 + e["n2o"] * GWP_N2O

def calc_fuel_emission(activity_val, unit, ef_key):
    """
    Base Year GHG Inventory pattern:
      tCO2e = activity_tj * (CO2 + CH4_tCO2e + N2O_tCO2e)
    For electricity: activity_MWh * EF_t/MWh
    """
    e = EF[ef_key]
    if e["unit"] == "t/MWh":
        return activity_val * e["co2"]
    else:
        # already in TJ
        return activity_val * (e["co2"] + e["ch4"] * GWP_CH4 + e["n2o"] * GWP_N2O)


def fuel_to_tj(fuel_type, amount, input_unit):
    """Convert fuel quantity to TJ using ListsAndTables conversion factors."""
    c = FUEL_CONV.get(fuel_type, {})
    if input_unit == "kl" and "kl_to_tj" in c:
        return amount * c["kl_to_tj"]
    if input_unit == "tonne" and "t_to_tj" in c:
        return amount * c["t_to_tj"]
    if input_unit == "tj":
        return amount
    return 0.0


# ─── BUILDINGS & ENERGY ───────────────────────────────────────────────────────
def calc_buildings(d):
    """
    Mirrors: Base Year GHG Inventory → Building and Energy Sector
    Form field names (from index.html):
      res_elec (MWh), res_lpg (t), res_firewood (t), res_kero (kL), res_png (t), res_coal (t), res_dg (kL)
      com_elec (MWh), com_lpg (t), com_png (t), com_firewood (t), com_kero (kL)
      ins_elec (MWh), ins_lpg (t)
      ind_elec (MWh), ind_lpg (t), ind_coal (t), ind_diesel (kL), ind_natgas (t)
      egen_coal (TJ), egen_natgas (TJ), egen_diesel (TJ)
    All solid/liquid fuels converted to TJ before applying EF (t/TJ).
    """
    subs = {}
 
    def emit_mwh(mwh, ef_key):
        """Electricity: MWh * EF(t/MWh) -> tCO2e"""
        return float(mwh or 0) * EF[ef_key]["co2"]
 
    def emit_tj(tj, ef_key):
        """Fuel already in TJ -> tCO2e (with GWP)"""
        e = EF[ef_key]
        return float(tj or 0) * (e["co2"] + e["ch4"] * GWP_CH4 + e["n2o"] * GWP_N2O)
 
    def emit_t(tonnes, fuel_type, ef_key):
        """tonnes → TJ → tCO2e"""
        tj = fuel_to_tj(fuel_type, float(tonnes or 0), "tonne")
        return emit_tj(tj, ef_key)
 
    def emit_kl(kl, fuel_type, ef_key):
        """kL → TJ → tCO2e"""
        tj = fuel_to_tj(fuel_type, float(kl or 0), "kl")
        return emit_tj(tj, ef_key)
 
    g = d.get  # shorthand
 
    subs["Residential"] = (
        emit_mwh(g("res_elec",     0), "Res_Electricity") +
        emit_t  (g("res_lpg",      0), "LPG",      "Res_LPG")      +
        emit_t  (g("res_firewood", 0), "Firewood", "Res_Firewood")  +
        emit_t  (g("res_png",      0), "PNG",      "Res_PNG")       +
        emit_t  (g("res_coal",     0), "Coal",     "Res_Coal")      +
        emit_kl (g("res_kero",     0), "Kerosene", "Res_Kerosene")  +
        emit_kl (g("res_dg",       0), "Diesel",   "Res_Diesel_Genset")
    )
 
    subs["Commercial"] = (
        emit_mwh(g("com_elec",     0), "Com_Electricity") +
        emit_t  (g("com_lpg",      0), "LPG",      "Com_LPG")      +
        emit_t  (g("com_png",      0), "PNG",      "Com_PNG")       +
        emit_t  (g("com_firewood", 0), "Firewood", "Com_Firewood")  +
        emit_kl (g("com_kero",     0), "Kerosene", "Com_Kerosene")
    )
 
    subs["Public & Institutional"] = (
        emit_mwh(g("ins_elec", 0), "Ins_Electricity") +
        emit_t  (g("ins_lpg",  0), "LPG", "Ins_LPG")
    )
 
    subs["Industrial"] = (
        emit_mwh(g("ind_elec",   0), "Ind_Electricity") +
        emit_t  (g("ind_lpg",    0), "LPG",     "Ind_LPG")    +
        emit_t  (g("ind_coal",   0), "Coal",    "Ind_Coal")   +
        emit_kl (g("ind_diesel", 0), "Diesel",  "Ind_Diesel") +
        emit_t  (g("ind_natgas", 0), "NatGas",  "Ind_NatGas")
    )
 
    # Energy Generation — form sends TJ directly (egen_coal, egen_natgas, egen_diesel)
    subs["Energy Generation"] = (
        emit_tj(g("egen_coal",   0), "EGen_Coal")    +
        emit_tj(g("egen_natgas", 0), "EGen_NatGas")  +
        emit_tj(g("egen_diesel", 0), "EGen_Diesel")
    )

    # ─── Expanded questionnaire: free-form fuel rows ────────────────────────
    def _sum_extra(prefix, target_subsector):
        total = 0.0
        for n in range(30):
            fuel = g(f"{prefix}_extra_fuel_{n}", "")
            val  = g(f"{prefix}_extra_value_{n}", "")
            unit = g(f"{prefix}_extra_unit_{n}", "")
            if fuel and val:
                total += emit_for_fuel(val, unit, fuel)
        if total:
            subs[target_subsector] = subs.get(target_subsector, 0.0) + total

    _sum_extra("res",  "Residential")
    _sum_extra("com",  "Commercial")
    _sum_extra("ins",  "Public & Institutional")

    # Manufacturing Industries & Construction (new tab)
    mfg_total = 0.0
    for n in range(60):
        fuel = g(f"mfg_row_{n}_fuel", "")
        val  = g(f"mfg_row_{n}_value", "")
        unit = g(f"mfg_row_{n}_unit", "")
        if fuel and val:
            mfg_total += emit_for_fuel(val, unit, fuel)
    if mfg_total:
        subs["Manufacturing Industries"] = mfg_total

    # Energy Industries (new tab)
    eind_total = 0.0
    for n in range(40):
        fuel = g(f"eind_row_{n}_fuel", "")
        val  = g(f"eind_row_{n}_value", "")
        unit = g(f"eind_row_{n}_unit", "")
        if fuel and val:
            eind_total += emit_for_fuel(val, unit, fuel)
    if eind_total:
        subs["Energy Generation"] = subs.get("Energy Generation", 0.0) + eind_total

    # Fugitive Emissions (new tab) — coal mining vents methane in m³/tonne
    fug_total = 0.0
    for n in range(40):
        sector = g(f"fug_row_{n}_sector", "")
        val    = g(f"fug_row_{n}_value", "")
        rate   = g(f"fug_row_{n}_rate", "")
        if sector and val:
            try:
                v = float(val or 0)
                r = float(rate or 0)
                ch4_t = v * r * 0.717e-3  # CH4 density 0.717 kg/m³
                fug_total += ch4_t * GWP_CH4
            except (TypeError, ValueError):
                pass
    if fug_total:
        subs["Fugitive Emissions"] = fug_total

    # Renewables (offsets) — MWh of clean generation, attached for dashboard
    ren_mwh = 0.0
    for n in range(30):
        val = g(f"ren_row_{n}_value", "")
        if val:
            try: ren_mwh += float(val or 0)
            except (TypeError, ValueError): pass
    if ren_mwh:
        subs["Renewables Offset (MWh)"] = ren_mwh

    return subs


# ─── TRANSPORT ────────────────────────────────────────────────────────────────
def calc_transport(d):
    """
    Mirrors D. Transportation sheet.
    Option 1 = Fuel Sales Approach (user enters fuel volumes by mode)
    Option 2 = VKT Approach (user enters vehicle counts & km/year)
    Form field 'trans_option' = '1' or '2' (default '1')
    """
    opt = str(d.get('trans_option', '1')).strip()
    if opt == '2':
        return _calc_transport_vkt(d)
    return _calc_transport_fuel_sales(d)
 
 
def _calc_transport_fuel_sales(d):
    """Option 1: Fuel Sales — matches sheet rows B4:B8, on-road table"""
 
    def emit(kl=0, t=0, mwh=0, kg=0, ef_key='Trans_Petrol'):
        e = EF[ef_key]
        if e['unit'] == 't/MWh': return mwh * e['co2']
        tj = 0.0
        if kl > 0: tj = kl * EF[ef_key].get('conv_kl', 0)
        if t  > 0: tj = t  * FUEL_CONV.get(ef_key.split('_')[-1], {}).get('t_to_tj', 0)
        if kg > 0: tj = kg * FUEL_CONV.get('Hydrogen', {}).get('kg_to_tj', 0.12)
        total_ef = e['co2'] + e['ch4'] * GWP_CH4 + e['n2o'] * GWP_N2O
        return tj * total_ef
 
    # ── ON ROAD ──────────────────────────────────────────────────────
    on_road = (
        emit(kl=float(d.get('t_pet',  0) or 0), ef_key='Trans_Petrol')     +
        emit(kl=float(d.get('t_die',  0) or 0), ef_key='Trans_Diesel')     +
        emit(t= float(d.get('t_cng',  0) or 0), ef_key='Trans_CNG')        +
        emit(t= float(d.get('t_alpg', 0) or 0), ef_key='Trans_AutoLPG')    +
        emit(t= float(d.get('t_lng',  0) or 0), ef_key='Trans_LNG')        +
        emit(kg=float(d.get('t_h2',   0) or 0), ef_key='Trans_Hydrogen')   +
        emit(kl=float(d.get('t_lub',  0) or 0), ef_key='Trans_Lubricants')
    )
    on_road += float(d.get('t_elec', 0) or 0) * EF['Trans_Electricity']['co2']  # EV MWh
 
    # ── RAILWAY ──────────────────────────────────────────────────────
    railway = (
        emit(kl=float(d.get('r_die',  0) or 0), ef_key='Trans_Railway_Die') +
        float(d.get('r_elec', 0) or 0) * EF['Trans_Railway_Ele']['co2']
    )
 
    # ── WATERBORNE NAVIGATION ────────────────────────────────────────
    water = (
        emit(kl=float(d.get('w_die', 0) or 0), ef_key='Trans_Water_Die') +
        emit(kl=float(d.get('w_pet', 0) or 0), ef_key='Trans_Water_Pet')
    )
 
    # ── AVIATION ─────────────────────────────────────────────────────
    aviation = (
        emit(kl=float(d.get('av_gas', 0) or 0), ef_key='Trans_AvGasoline') +
        emit(kl=float(d.get('av_jet', 0) or 0), ef_key='Trans_JetKerosene')
    )
 
    return {
        'On Road':                on_road,
        'Railway':                railway,
        'Water Borne Navigation': water,
        'Aviation':               aviation,
    }
 
def _calc_transport_vkt(d):
    """
    Option 2: Vehicular Kilometre Travel (VKT)
    Mirrors D. Transportation VKT sub-table (vehicle count × km/yr × EF_km)
    EF_km (tCO2e/km) = (fuel_consumption_L_per_100km / 100) * (fuel_density kg/L)
                       * (TJ_per_tonne) * (CO2+CH4+N2O per TJ)
 
    Form fields expected:
      vkt_motorcycle_count, vkt_motorcycle_fuel, vkt_motorcycle_km
      vkt_car_count, vkt_car_fuel, vkt_car_km
      ... (one set per vehicle type)
    """
    # VKT emission factors (tCO2e per vehicle-km) — IPCC defaults
    VKT_EF = {
        'Motorcycle_Petrol':   0.000082,
        'Car_Petrol':          0.000192,
        'Car_Diesel':          0.000171,
        'Car_CNG':             0.000135,
        'Car_EV':              0.000097,  # grid EF applied
        'Bus_Diesel':          0.000890,
        'Bus_CNG':             0.000780,
        'HDTruck_Diesel':      0.001250,
        'LDTruck_Diesel':      0.000210,
    }
 
    on_road = 0.0
    veh_types = ['motorcycle', 'car', 'taxi', 'bus', 'ldtruck', 'mdtruck', 'hdtruck', 'metro']
    for vt in veh_types:
        count = float(d.get(f'vkt_{vt}_count', 0) or 0)
        km    = float(d.get(f'vkt_{vt}_km',    0) or 0)
        fuel  = d.get(f'vkt_{vt}_fuel', 'Petrol')
        key   = f'{vt.capitalize()}_{fuel}'
        ef    = VKT_EF.get(key, VKT_EF.get(f'Car_{fuel}', 0.000192))
        on_road += count * km * ef
 
    # Railway (electric only in VKT mode)
    railway = float(d.get('r_elec', 0) or 0) * EF['Trans_Railway_Ele']['co2']
 
    return {
        'On Road':               on_road,
        'Railway':               railway,
        'Water Borne Navigation': 0.0,
        'Aviation':              0.0,
    }
    
# ─── SOLID WASTE ──────────────────────────────────────────────────────────────
def calc_solid_waste(d):
    """
    Mirrors: E. Solid Waste → Landfill CH4 (IPCC FOD method)
    CH4_landfill = MSW_landfill × DOC_weighted × DOCF × F × 16/12 × (1-OX) × MCF
    Biogas/Composting CH4 from organic fraction
    """
    sw_tot = float(d.get("sw_tot", 0) or 0)   # tonne/day total MSW
    if sw_tot <= 0:
        return {"Solid Waste Disposal": 0.0, "Organic Waste Treatment": 0.0}
 
    sw_tpa = sw_tot * 365  # tonne/year
 
    # Waste fractions (default from E. Solid Waste sheet)
    f_food  = float(d.get("sw_food_frac_pct", d.get("sw_food_frac", 72.6)) or 72.6) / 100.0
    f_paper = float(d.get("sw_paper_frac_pct", d.get("sw_paper_frac", 3.5))  or 3.5)  / 100.0
    f_other = max(0.0, 1.0 - f_food - f_paper)
 
    # Landfill fraction
    lfm = float(d.get("sw_lfm_pct", d.get("sw_lfm", 85.0)) or 85.0) / 100.0
    lfu = float(d.get("sw_lfu_pct", d.get("sw_lfu", 0.0))  or 0.0)  / 100.0
    sw_landfill_tpa = sw_tpa * (lfm + lfu)
 
    # Weighted DOC
    doc_w = (f_food * SW_DOC["food"] + f_paper * SW_DOC["paper"] +
             f_other * SW_DOC.get("rubber", 0.0))
    doc_w = max(doc_w, 0.05)
 
    # IPCC FOD constants (from E. Solid Waste sheet R89-94)
    docf = 0.6    # fraction of DOC ultimately decomposed
    f_ch4 = 0.5   # fraction of CH4 in landfill gas
    ox = 0.1      # oxidation factor
    mcf_managed = 1.0  # managed landfill
 
    # CH4 emissions (tCH4)
    ch4_gen = sw_landfill_tpa * doc_w * docf * f_ch4 * (16.0/12.0) * mcf_managed
    ch4_emit = ch4_gen * (1.0 - ox)  # subtract oxidised
 
    # Collection / recovery (default 0 unless input)
    collection_eff = float(d.get("sw_gas_collection_pct", d.get("sw_gas_collection", 0)) or 0) / 100.0
    ch4_recovered  = ch4_emit * collection_eff
    ch4_net = ch4_emit - ch4_recovered
 
    tco2e_landfill = ch4_net * GWP_CH4
 
    # Organic waste treatment (incineration / composting)
    inc_frac = float(d.get("sw_inc_pct", d.get("sw_inc", 0.4)) or 0.4) / 100.0
    inc_tpa  = sw_tpa * inc_frac
    # Incineration: IPCC default EF 91.7 tCO2/TJ for non-biomass MSW
    # Using simplified: 0.5 tCO2e/t for non-biomass fraction
    tco2e_inc = inc_tpa * 0.5
 
    return {
        "Solid Waste Disposal":    tco2e_landfill,
        "Organic Waste Treatment": tco2e_inc,
    }


# ─── WASTEWATER ───────────────────────────────────────────────────────────────
def calc_wastewater(d):
    """
    Mirrors: F. Waste water Emission
    CH4 = BOD × B0 × MCF_weighted
    N2O from effluent discharge (protein-based)
    """
    population = float(d.get("population", 0) or 0)
    lpcd       = float(d.get("ww_lpcd", 135) or 135)   # L/person/day
    bod_pc     = float(d.get("ww_bod", 34) or 34)       # g BOD/person/day (sheet default 34)
    tn_pc      = float(d.get("ww_tn", 0.026) or 0.026)  # kg N/person/day (Table default)
 
    # BOD total (kg/year)
    bod_total = population * bod_pc / 1000.0 * 365  # kg/year
    # Total Nitrogen
    tn_total  = population * tn_pc * 365             # kg/year
    # Industrial co-discharge factor (from sheet: 1.056521739)
    co_factor = float(d.get("ww_co_factor", 1.0565) or 1.0565)
    bod_total *= co_factor
    tn_total  *= co_factor
 
    # Treatment fractions — form sends ww_aer_pct etc. (0-100), convert to fraction
    f_aer_not_well = float(d.get("ww_aer_pct",  d.get("ww_aer",  17.0)) or 17.0) / 100.0
    f_anaerobic_r  = float(d.get("ww_uasb_pct", d.get("ww_uasb",  9.0)) or  9.0) / 100.0
    f_septic       = float(d.get("ww_sep_pct",  d.get("ww_sep",   0.0)) or  0.0) / 100.0
    f_open         = float(d.get("ww_open_pct", d.get("ww_open",  6.0)) or  6.0) / 100.0
    f_untreated    = max(0.0, 1.0 - f_aer_not_well - f_anaerobic_r - f_septic - f_open)
 
    # B0 = 0.6 kg CH4/kg BOD (max capacity, from sheet R62)
    b0 = 0.6
 
    # CH4 by treatment type (kg CH4/yr)
    ch4_aer = bod_total * f_aer_not_well * b0 * WW_MCF["aerobic_ponds"]      # 0.0 (well managed)
    ch4_aer_nw = bod_total * f_aer_not_well * b0 * 0.3   # not well managed MCF=0.3
    ch4_anaer = bod_total * f_anaerobic_r * b0 * WW_MCF["anaerobic_reactor"]
    ch4_septic= bod_total * f_septic * b0 * WW_MCF["septic"]
    ch4_open  = bod_total * f_open   * b0 * WW_MCF["open_discharge"]
    ch4_unt   = bod_total * f_untreated * b0 * 0.1
 
    ch4_total_kg = ch4_aer_nw + ch4_anaer + ch4_septic + ch4_open + ch4_unt
    ch4_total_t  = ch4_total_kg / 1000.0
    tco2e_ch4    = ch4_total_t * GWP_CH4
 
    # N2O from effluent (IPCC 2019 Eq 6.9)
    # N2O-N from effluent = TN * EF_effluent (0.005 kg N2O-N/kg N)
    n2o_n = tn_total * 0.005   # kg N2O-N/yr
    n2o_t = n2o_n * (44.0/28.0) / 1000.0  # tN2O/yr
    tco2e_n2o = n2o_t * GWP_N2O
 
    return {"Waste water": tco2e_ch4 + tco2e_n2o}

# ─── AFOLU ────────────────────────────────────────────────────────────────────
def calc_afolu(d):
    """
    Mirrors: H. AFOLU sheet
    Enteric fermentation + Manure management CH4/N2O
    Land use change (simplified)
    """
    total_enteric = 0.0
    total_manure  = 0.0
 
    # Form field names match AFOLU EF keys directly
    livestock_map = {
        "dairy_cow_indigenous": ("dairy_cow_indigenous", "dairy_cow_indigenous"),
        "dairy_cow_crossbred":  ("dairy_cow_crossbred",  "dairy_cow_crossbred"),
        "nondairy_cow_adult":   ("nondairy_cow_adult",   "nondairy_cow_adult"),
        "dairy_buffalo":        ("dairy_buffalo",        "dairy_buffalo"),
        "sheep":                ("sheep",                "sheep"),
        "goat":                 ("goat",                 "goat"),
        "swine":                ("swine",                "swine"),
        "poultry":              ("poultry",              "poultry"),
    }
    for ltype, (form_key, ef_key) in livestock_map.items():
        heads = float(d.get(form_key, 0) or 0)
        if heads > 0:
            # Enteric CH4
            ef_e = AFOLU_ENTERIC.get(ef_key, 0)
            ch4_e_t = heads * ef_e / 1000.0  # kg/head/yr → t/yr
            total_enteric += ch4_e_t * GWP_CH4
            # Manure CH4
            ef_m = AFOLU_MANURE_CH4.get(ef_key, 0)
            ch4_m_t = heads * ef_m / 1000.0
            total_manure += ch4_m_t * GWP_CH4
 
    # Wetland rice CH4 (simple: area × EF)
    wet_ha = float(d.get("paddy_ha", d.get("af_wet", 0)) or 0)
    # IPCC default EF: 1.3 kgCH4/ha/day, 120 day season
    ch4_wet_t = wet_ha * 1.3 * 120 / 1000.0 if wet_ha > 0 else 0.0
    total_wetland = ch4_wet_t * GWP_CH4
 
    # Forestland CO2 sequestration (negative emissions)
    forest_ha = float(d.get("green_ha", d.get("af_fd", 0)) or 0)
    # Average growth 5 tCO2/ha/yr for tropical moist (IPCC Tier 1)
    seq_forest = -forest_ha * 5.0 if forest_ha > 0 else 0.0
 
    # Grassland / managed land (simplified)
    grass_ha  = float(d.get("af_fm", 0) or 0)  # no separate grass field in form
    other_ha  = float(d.get("af_fo", 0) or 0)
    seq_land   = -(grass_ha + other_ha) * 1.5 if (grass_ha + other_ha) > 0 else 0.0
 
    return {
        "Live Stock":          total_enteric + total_manure + total_wetland,
        "Land Management":     seq_forest + seq_land,
        "Aggregate Sources":   0.0,
    }

# ─── IPPU ─────────────────────────────────────────────────────────────────────
def calc_ippu(d):
    """
    Mirrors: G. IPPU sheet
    Mineral + Chemical + Metal industries
    """
    subs = {}
 
    # Cement (clinker-based, IPCC Eq 2.2 Tier 2)
    # Form field names: cement_clinker, lime_high_ca, steel_bof, steel_eaf, ammonia, glass
    clinker_t = float(d.get("cement_clinker", d.get("ip_clink", 0)) or 0)
    subs["Mineral Industry"] = (
        clinker_t * IPPU_EF["cement_clinker"] +
        float(d.get("lime_high_ca", d.get("ip_lime", 0)) or 0) * IPPU_EF["lime_high_ca"]
    )
 
    # Chemical (ammonia)
    nh3_t    = float(d.get("ammonia", d.get("ip_nh3", 0)) or 0)
    hno3_t   = float(d.get("ip_hno3", 0) or 0)
    n2o_hno3 = hno3_t * IPPU_EF["hno3_n2o"] / 1000.0 * GWP_N2O
    subs["Chemical Industry"] = nh3_t * IPPU_EF["ammonia"] + n2o_hno3
 
    # Metal (steel BOF/EAF)
    bof_t = float(d.get("steel_bof", d.get("ip_bof", 0)) or 0)
    eaf_t = float(d.get("steel_eaf", d.get("ip_eaf", 0)) or 0)
    subs["Metal Industry"] = bof_t * IPPU_EF["steel_bof"] + eaf_t * IPPU_EF["steel_eaf"]
 
    subs["Non-Energy Products"] = float(d.get("glass", 0) or 0) * IPPU_EF["glass_ef"]
    subs["Ozone Depleting Substances"]      = 0.0
    subs["Other Product Manufacture and Use"]= 0.0
 
    return subs

# ─── BAU PROJECTIONS ─────────────────────────────────────────────────────────
def calc_bau(base_emissions_by_sector, d, year):
    """
    Mirrors: BAU Scenario sheet — per-period growth using Population & GDP Growth Factor.
    Each planning period (base→interim1, interim1→interim2, interim2→target) uses
    its own compound growth factor derived from pop_growth and gdp_growth inputs.
 
    Growth Factor (per period) = (1 + pop_rate/100) * (1 + gdp_rate/100) - 1
    Projection compounds across periods up to the requested year.
 
    Special case: Railway stays flat from 2040 onwards (as in sheet R19)
    """
    base_year   = int(d.get("base_year",   2025))
    interim1    = int(d.get("interim1",    2030))
    interim2    = int(d.get("interim2",    2040))
    target_year = int(d.get("target_year", 2050))
 
    def period_factor(pop_key, gdp_key, fallback_rate=0.03):
        """Compute compound growth factor for a period from pop + gdp rates."""
        pop_rate = float(d.get(pop_key, d.get("pop_growth_rate", fallback_rate * 100)) or (fallback_rate * 100)) / 100.0
        gdp_rate = float(d.get(gdp_key, d.get("gdp_growth_rate", fallback_rate * 100)) or (fallback_rate * 100)) / 100.0
        return (1 + pop_rate) * (1 + gdp_rate) - 1
 
    # Per-period annual growth rates
    r_base    = period_factor("base_pop_growth",    "base_gdp_growth",    0.02)
    r_interim1= period_factor("interim1_pop_growth","interim1_gdp_growth", 0.03)
    r_interim2= period_factor("interim2_pop_growth","interim2_gdp_growth", 0.02)
    r_target  = period_factor("target_pop_growth",  "target_gdp_growth",   0.03)
 
    def compound_factor(yr):
        """Compound growth factor from base_year to yr, using per-period rates."""
        if yr <= base_year:
            return 1.0
 
        # Period 1: base → interim1
        p1_end   = min(yr, interim1)
        p1_years = max(0, p1_end - base_year)
        f        = (1 + r_base) ** p1_years
 
        if yr <= interim1:
            return f
 
        # Period 2: interim1 → interim2
        p2_end   = min(yr, interim2)
        p2_years = max(0, p2_end - interim1)
        f       *= (1 + r_interim1) ** p2_years
 
        if yr <= interim2:
            return f
 
        # Period 3: interim2 → target
        p3_years = max(0, yr - interim2)
        f       *= (1 + r_interim2) ** p3_years
 
        return f
 
    # Special factor for Railway (flat from 2040)
    railway_flat_year = 2040
    factor = compound_factor(year)
 
    projected = {}
    for sector, subsectors in base_emissions_by_sector.items():
        projected[sector] = {}
        for sub, val in subsectors.items():
            if sector == "Transport" and sub == "Railway" and year >= railway_flat_year:
                factor_sub = compound_factor(railway_flat_year)
            else:
                factor_sub = factor
            projected[sector][sub] = val * factor_sub
    return projected

# ─── TARGET SETTING ──────────────────────────────────────────────────────────
def calc_targets(bau_totals, d):
    """
    Mirrors: Target Setting sheet
    target_val = BAU - (target_pct * BAU_base)
    """
    base_year   = int(d.get("base_year", 2025))
    target_year = int(d.get("target_year", 2050))
    interim1    = int(d.get("interim1", 2030))
    interim2    = int(d.get("interim2", 2040))
    target_pct  = float(d.get("target_pct", 0.45) or 0.45)

    base_total = bau_totals.get(base_year, 0)

    targets = {}
    for yr, bau in bau_totals.items():
        if yr == base_year:
            targets[yr] = base_total
        else:
            # Linear % reduction scaling to target_year
            frac = min(1.0, (yr - base_year) / max(1, (target_year - base_year)))
            reduction = base_total * target_pct * frac
            targets[yr] = bau - reduction
    return targets


# ─── EMISSION REDUCTION GRAPH — E&P and High Ambition ────────────────────────
def calc_scenarios(base_emissions_by_sector, bau_by_year, d):
    """
    Mirrors: Emission Reduction- Graph sheet
    For each sector×subsector×year:
      EP_reduction  = BAU × ep_pct  (user slider)
      HA_reduction  = BAU × ha_pct  (user slider)
      EP_emission   = BAU - EP_reduction
      HA_emission   = BAU - HA_reduction
    Then aggregate.
    """
    years = sorted(bau_by_year.keys())

    ep_total = {}
    ha_total = {}

    SUBSECTOR_KEYS = {
        "Residential":            "Residential",
        "Commercial":             "Commercial",
        "Public & Institutional": "Public___Institutional",
        "Industrial":             "Industrial",
        "Energy Generation":      "Energy_Generation",
        "On Road":                "On_Road",
        "Railway":                "Railway",
        "Water Borne Navigation": "Water_Borne_Navigation",
        "Aviation":               "Aviation",
        "Solid Waste Disposal":   "Solid_Waste_Disposal",
        "Organic Waste Treatment":"Organic_Waste_Treatment",
        "Waste water":            "Waste_water",
        "Live Stock":             "Live_Stock",
        "Land Management":        "Land_Management",
        "Aggregate Sources":      "Aggregate_Sources",
        "Mineral Industry":       "Mineral_Industry",
        "Chemical Industry":      "Chemical_Industry",
        "Metal Industry":         "Metal_Industry",
    }

    for yr in years:
        bau_yr = bau_by_year[yr]
        ep_yr  = 0.0
        ha_yr  = 0.0
        for sector, subsectors in bau_yr.items():
            for sub, bau_val in subsectors.items():
                key = SUBSECTOR_KEYS.get(sub, sub.replace(" ","_").replace("&","").replace("/","_"))
                # Get slider values (default from Excel sheet E&P/HA defaults)
                ep_pct_key = f"ep_pct_{key}"
                ha_pct_key = f"ha_pct_{key}"
                ep_pct = float(d.get(ep_pct_key, _ep_default(sub)) or _ep_default(sub)) / 100.0
                ha_pct = float(d.get(ha_pct_key, _ha_default(sub)) or _ha_default(sub)) / 100.0
                ep_yr += bau_val * (1.0 - ep_pct)
                ha_yr += bau_val * (1.0 - ha_pct)
        ep_total[yr] = ep_yr
        ha_total[yr] = ha_yr

    return ep_total, ha_total


def _ep_default(subsector):
    """Default E&P reduction % from Emission Reduction-Graph sheet (col E)."""
    EP_DEFAULTS = {
        "Residential": 10, "Commercial": 5, "Public & Institutional": 5,
        "Industrial": 5, "Energy Generation": 0, "On Road": 5, "Railway": 5,
        "Water Borne Navigation": 5, "Aviation": 10,
        "Solid Waste Disposal": 5, "Organic Waste Treatment": 5,
        "Waste water": 5, "Live Stock": 5, "Land Management": 5,
        "Mineral Industry": 5, "Chemical Industry": 5, "Metal Industry": 5,
    }
    return EP_DEFAULTS.get(subsector, 5)


def _ha_default(subsector):
    """Default High Ambition reduction % from Emission Reduction-Graph sheet (col H)."""
    HA_DEFAULTS = {
        "Residential": 30, "Commercial": 30, "Public & Institutional": 30,
        "Industrial": 30, "Energy Generation": 20, "On Road": 35, "Railway": 20,
        "Water Borne Navigation": 20, "Aviation": 30,
        "Solid Waste Disposal": 40, "Organic Waste Treatment": 40,
        "Waste water": 40, "Live Stock": 15, "Land Management": 15,
        "Mineral Industry": 20, "Chemical Industry": 20, "Metal Industry": 20,
    }
    return HA_DEFAULTS.get(subsector, 20)


# ─── MITIGATION BUDGET (Strategies & Cost) ────────────────────────────────────
def calc_mitigation_budget(base_by_sector, bau_by_year, ha_by_year, d):
    """
    Mirrors: Stratergies & Cost + Dashboard- Scenario Comparison
    For each sector: GHG reduced = BAU_targetyear - HA_targetyear
    Investment = GHG_reduced * cost_per_tonne (sector-specific)
    """
    target_year = int(d.get("target_year", 2050))
    bau_ty = bau_by_year.get(target_year, {})
    ha_ty  = ha_by_year.get(target_year, 0)

    # Sector-level BAU aggregates at target year
    sector_bau = {}
    for sector, subsectors in bau_ty.items():
        sector_bau[sector] = sum(subsectors.values())

    # HA total at target year
    bau_total_ty = sum(sector_bau.values())
    ha_total_ty  = ha_by_year.get(target_year, bau_total_ty)

    budget_rows = []
    total_reduced = 0.0
    total_inv = 0.0

    sector_map = {
        "Energy Sector":  ("Buildings", ABATEMENT_COST["Buildings"]),
        "Transport":      ("Transport", ABATEMENT_COST["Transport"]),
        "Waste":          ("Waste",     ABATEMENT_COST["Waste"]),
        "Wastewater":     ("Wastewater",ABATEMENT_COST["Wastewater"]),
        "AFOLU":          ("AFOLU",     ABATEMENT_COST["AFOLU"]),
        "IPPU":           ("IPPU",      ABATEMENT_COST["IPPU"]),
    }

    for sector, bau_val in sector_bau.items():
        display, cost_t = sector_map.get(sector, (sector, 2000))
        # Proportional HA reduction
        if bau_total_ty > 0:
            ha_val = ha_total_ty * (bau_val / bau_total_ty)
        else:
            ha_val = 0.0

        # User-defined reduction %
        red_pct_key = f"ha_pct_{display.replace(' ','_')}"
        red_pct = float(d.get(red_pct_key, 20) or 20) / 100.0
        reduced = bau_val * red_pct
        inv = reduced * cost_t / 1e7  # ₹ Crore

        total_reduced += reduced
        total_inv     += inv
        budget_rows.append({
            "Sector":               display,
            "BAU (t CO2e)":         round(bau_val),
            "Reduction %":          f"{red_pct*100:.0f}%",
            "GHG Reduced (t CO2e)": round(reduced),
            "Investment (Crore)":   f"₹{inv:,.1f}",
        })

    budget_rows.append({
        "Sector":               "TOTAL",
        "BAU (t CO2e)":         round(bau_total_ty),
        "Reduction %":          f"{(total_reduced/bau_total_ty*100):.1f}%" if bau_total_ty else "—",
        "GHG Reduced (t CO2e)": round(total_reduced),
        "Investment (Crore)":   f"₹{total_inv:,.1f}",
    })

    return budget_rows, total_inv


# ─── MILESTONE TABLE (Target Setting) ─────────────────────────────────────────
def calc_milestones(bau_totals, ep_totals, ha_totals, targets, d):
    """
    Mirrors: Target Setting + Emission Reduction-Graph R99-103
    """
    rows = []
    base_year = int(d.get("base_year", 2025))
    years = sorted(bau_totals.keys())
    for yr in years:
        bau = bau_totals[yr]
        tgt = targets.get(yr, bau)
        ha  = ha_totals.get(yr, bau)
        ep  = ep_totals.get(yr, bau)
        base= bau_totals[base_year]
        req_pct = (bau - tgt) / bau * 100 if bau > 0 else 0
        ach_pct = (bau - ha)  / bau * 100 if bau > 0 else 0
        status  = "On Track" if ha <= tgt else "Gap"
        rows.append({
            "year":         yr,
            "bau":          round(bau / 1e6, 2),
            "target":       round(tgt / 1e6, 2),
            "ep":           round(ep  / 1e6, 2),
            "ha":           round(ha  / 1e6, 2),
            "required_pct": f"{req_pct:.1f}%",
            "achieved_pct": f"{ach_pct:.1f}%",
            "status":       status,
        })
    return rows


# ─── PLOTLY CHARTS ────────────────────────────────────────────────────────────
NAVY  = "#1a2744"
TEAL  = "#00b4a6"
AMBER = "#f59e0b"
GREEN = "#10b981"
RED   = "#ef4444"
BLUE  = "#3b82f6"
CHART_FONT = dict(family="DM Sans, sans-serif", size=12, color="#444")


def make_trajectory_chart(bau_totals, ep_totals, ha_totals, targets, years):
    """Mirrors Emission Reduction-Graph dashboard chart (rows 43-47, 99-103)."""
    fig = go.Figure()

    ys = sorted(years)
    scale = 1e6  # tCO2e → MtCO2e

    fig.add_trace(go.Scatter(
        x=ys, y=[bau_totals[y]/scale for y in ys],
        name="Reference (BAU)", mode="lines+markers",
        line=dict(color=NAVY, width=2.5, dash="solid"),
        marker=dict(size=6)
    ))
    fig.add_trace(go.Scatter(
        x=ys, y=[ep_totals.get(y, bau_totals[y])/scale for y in ys],
        name="Existing & Planned (E&P)", mode="lines+markers",
        line=dict(color=BLUE, width=2, dash="dot"),
        marker=dict(size=6)
    ))
    fig.add_trace(go.Scatter(
        x=ys, y=[ha_totals.get(y, bau_totals[y])/scale for y in ys],
        name="High Ambition", mode="lines+markers",
        line=dict(color=TEAL, width=2.5),
        marker=dict(size=7)
    ))
    fig.add_trace(go.Scatter(
        x=ys, y=[targets.get(y, bau_totals[y])/scale for y in ys],
        name="Target Pathway", mode="lines+markers",
        line=dict(color=GREEN, width=2, dash="dash"),
        marker=dict(symbol="diamond", size=8)
    ))

    # Shaded gap area between BAU and HA
    ys_r = list(reversed(ys))
    fig.add_trace(go.Scatter(
        x=ys + ys_r,
        y=[bau_totals[y]/scale for y in ys] + [ha_totals.get(y, bau_totals[y])/scale for y in ys_r],
        fill="toself", fillcolor="rgba(0,180,166,0.10)",
        line=dict(color="rgba(0,0,0,0)"), showlegend=False, hoverinfo="skip"
    ))

    fig.update_layout(
        title=dict(text="GHG Emission Trajectory (Mt CO₂e)", font=dict(size=14)),
        xaxis=dict(title="Year", gridcolor="#f0f0f0"),
        yaxis=dict(title="Mt CO₂e", gridcolor="#f0f0f0"),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        plot_bgcolor="white", paper_bgcolor="white",
        font=CHART_FONT, margin=dict(l=50, r=20, t=60, b=40),
        hovermode="x unified"
    )
    return json.loads(fig.to_json())


def make_pie_chart(base_by_sector):
    """Mirrors Dashboard- BAU-City sector breakdown pie chart."""
    labels, values = [], []
    sector_colors = {
        "Energy Sector":  NAVY,
        "Transport":      TEAL,
        "Waste":          AMBER,
        "Wastewater":     BLUE,
        "AFOLU":          GREEN,
        "IPPU":           "#8b5cf6",
    }
    for sector, subs in base_by_sector.items():
        total = sum(subs.values())
        if total > 0:
            labels.append(sector)
            values.append(total)

    fig = go.Figure(go.Pie(
        labels=labels, values=values,
        hole=0.42,
        marker=dict(colors=[sector_colors.get(l, "#aaa") for l in labels]),
        textinfo="label+percent",
        textfont=dict(size=11),
        hovertemplate="%{label}<br>%{value:,.0f} tCO₂e<br>%{percent}<extra></extra>"
    ))
    fig.update_layout(
        title=dict(text="Base Year Emissions by Sector", font=dict(size=14)),
        showlegend=False, font=CHART_FONT,
        plot_bgcolor="white", paper_bgcolor="white",
        margin=dict(l=10, r=10, t=50, b=10)
    )
    return json.loads(fig.to_json())


def make_bar_chart(bau_totals, ep_totals, ha_totals, targets):
    """Grouped bar chart — mirrors Dashboard- Scenario Comparison."""
    years = sorted(bau_totals.keys())
    scale = 1e6

    fig = go.Figure()
    fig.add_trace(go.Bar(name="BAU",          x=years,
                         y=[bau_totals[y]/scale for y in years], marker_color=NAVY))
    fig.add_trace(go.Bar(name="E&P",          x=years,
                         y=[ep_totals.get(y,0)/scale for y in years], marker_color=BLUE))
    fig.add_trace(go.Bar(name="High Ambition",x=years,
                         y=[ha_totals.get(y,0)/scale for y in years], marker_color=TEAL))
    fig.add_trace(go.Bar(name="Target",       x=years,
                         y=[targets.get(y,0)/scale for y in years],   marker_color=GREEN))

    fig.update_layout(
        barmode="group",
        title=dict(text="Scenario Comparison by Year", font=dict(size=14)),
        xaxis=dict(title="Year", type="category"),
        yaxis=dict(title="Mt CO₂e", gridcolor="#f0f0f0"),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        plot_bgcolor="white", paper_bgcolor="white",
        font=CHART_FONT, margin=dict(l=50, r=20, t=60, b=40),
    )
    return json.loads(fig.to_json())


def make_budget_chart(budget_rows):
    """Mitigation cost waterfall by sector."""
    rows = [r for r in budget_rows if r["Sector"] != "TOTAL"]
    sectors = [r["Sector"] for r in rows]
    reductions = [r["GHG Reduced (t CO2e)"] / 1e6 for r in rows]

    fig = go.Figure(go.Bar(
        x=sectors, y=reductions,
        marker_color=TEAL,
        text=[f"{v:.2f} Mt" for v in reductions],
        textposition="outside",
        hovertemplate="%{x}<br>Reduced: %{y:.3f} Mt CO₂e<extra></extra>"
    ))
    fig.update_layout(
        title=dict(text="GHG Reduction Potential by Sector (Target Year)", font=dict(size=14)),
        xaxis=dict(title="Sector"),
        yaxis=dict(title="Mt CO₂e Reduced", gridcolor="#f0f0f0"),
        plot_bgcolor="white", paper_bgcolor="white",
        font=CHART_FONT, margin=dict(l=50, r=20, t=60, b=80),
    )
    return json.loads(fig.to_json())


def make_subsector_bar(base_by_sector):
    """Stacked sub-sector bar."""
    sectors, totals = [], []
    for sector, subs in base_by_sector.items():
        t = sum(subs.values())
        if t > 0:
            sectors.append(sector)
            totals.append(t / 1e6)

    fig = go.Figure(go.Bar(
        x=sectors, y=totals,
        marker_color=[NAVY, TEAL, AMBER, BLUE, GREEN, "#8b5cf6"][:len(sectors)],
        text=[f"{v:.2f}" for v in totals],
        textposition="outside",
    ))
    fig.update_layout(
        title=dict(text="Base Year Emissions by Sector (Mt CO₂e)", font=dict(size=14)),
        xaxis=dict(title="Sector"),
        yaxis=dict(title="Mt CO₂e", gridcolor="#f0f0f0"),
        plot_bgcolor="white", paper_bgcolor="white",
        font=CHART_FONT, margin=dict(l=50, r=20, t=60, b=80),
    )
    return json.loads(fig.to_json())


# ═══════════════════════════════════════════════════════════════════════════════
# FLASK ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/")
@login_required
def index():
    return render_template("index.html", cities=INDIA_CITIES)


@app.route("/api/metadata")
@login_required
def api_metadata():
    """Return pre-populated secondary data for a location, mapped to form fields.

    Query params: ?state=<state>&district=<district>
    Only called when the user has turned on 'Use secondary data?'.
    """
    from data import metadata_lookup
    state    = (request.args.get("state") or "").strip()
    district = (request.args.get("district") or "").strip()
    if not metadata_lookup.available():
        return jsonify({"ok": False, "error": "Metadata dataset not installed.",
                        "fields": {}}), 200
    fields = metadata_lookup.lookup(state, district)
    return jsonify({
        "ok": True,
        "state": state,
        "district": district,
        "count": len(fields),
        "fields": fields,
    })


@app.route("/results")
@login_required
def results():
    return render_template("results.html")


@app.route("/target")
@login_required
def target_page():
    return render_template("target.html")


@app.route("/api/target-calc", methods=["POST"])
@login_required
def api_target_calc():
    """Interactive Target Setting tool — mirrors the Excel 'Target Setting' sheet.

    Excel logic (row 18-21 = Base / Interim1 / Interim2 / Target):
        Target(MtCO2eq) = BAU - IF(K15 = 2, TargetValue, BAU * TargetPct)

    K15 is the mode toggle:
        mode == "pct"   (K15 = 1)  → reduce BAU by a percentage per year
        mode == "value" (K15 = 2)  → subtract an absolute MtCO2e value per year

    'basis' mirrors the E12 dropdown ('BAU calculations are based on'):
        "population" | "gdp"  — selects which growth driver leads the BAU curve.
    """
    payload = request.get_json(force=True) or {}
    d = payload.get("inputs") or session.get("last_inputs") or {}
    if not d:
        # Fall back to the logged-in user's saved form so the tool works even on
        # a fresh page load before any /api/calculate has run this session.
        u = current_user()
        if u:
            saved = _load_user_form(u)
            d = (saved or {}).get("form", {})
    mode  = payload.get("mode", "pct")            # "pct" | "value"
    basis = payload.get("basis", "population")    # "population" | "gdp"
    # Per-year target settings from the tool: {year: {pct, value}}
    rows_in = payload.get("rows", {})

    base_year   = int(d.get("base_year",   2025))
    interim1    = int(d.get("interim1",    2030))
    interim2    = int(d.get("interim2",    2040))
    target_year = int(d.get("target_year", 2050))
    years = [base_year, interim1, interim2, target_year]

    # Recompute base-year emissions and BAU per year from the saved inputs.
    d_eff = dict(d)
    # The 'basis' dropdown decides whether pop or gdp growth leads the BAU.
    # We express that by zeroing the non-selected driver's contribution so the
    # BAU curve is driven by the chosen growth rate (matches the sheet's intent).
    if basis == "population":
        for k in ("base_gdp_growth", "interim1_gdp_growth",
                  "interim2_gdp_growth", "target_gdp_growth"):
            d_eff.setdefault(k, 0)
    elif basis == "gdp":
        for k in ("base_pop_growth", "interim1_pop_growth",
                  "interim2_pop_growth", "target_pop_growth"):
            d_eff.setdefault(k, 0)

    base_by_sector = {
        "Energy Sector": calc_buildings(d_eff),
        "Transport":     calc_transport(d_eff),
        "Waste":         calc_solid_waste(d_eff),
        "Wastewater":    calc_wastewater(d_eff),
        "AFOLU":         calc_afolu(d_eff),
        "IPPU":          calc_ippu(d_eff),
    }

    bau_totals = {}
    for yr in years:
        proj = calc_bau(base_by_sector, d_eff, yr)
        bau_totals[yr] = sum(sum(s.values()) for s in proj.values())

    # Build the per-year target table exactly like the sheet.
    table = []
    for i, yr in enumerate(years):
        bau_mt = bau_totals.get(yr, 0) / 1e6
        row_cfg = rows_in.get(str(yr), {})
        pct   = float(row_cfg.get("pct", 0) or 0)        # 0-100 (%)
        value = float(row_cfg.get("value", 0) or 0)      # absolute MtCO2e reduction

        if i == 0:
            # Base year — no reduction, target == BAU
            target_mt = bau_mt
            applied_pct = 0.0
            applied_val = 0.0
        elif mode == "value":
            target_mt = max(0.0, bau_mt - value)
            applied_val = value
            applied_pct = (value / bau_mt * 100) if bau_mt else 0.0
        else:  # "pct"
            target_mt = max(0.0, bau_mt - bau_mt * (pct / 100.0))
            applied_pct = pct
            applied_val = bau_mt * (pct / 100.0)

        reduction_mt = bau_mt - target_mt
        table.append({
            "year":         yr,
            "label":        ["Base Year", "Interim Year 1", "Interim Year 2", "Target Year"][i],
            "bau_mt":       round(bau_mt, 3),
            "target_pct":   round(applied_pct, 1),
            "target_value": round(applied_val, 3),
            "target_mt":    round(target_mt, 3),
            "reduction_mt": round(reduction_mt, 3),
        })

    return jsonify({
        "mode":  mode,
        "basis": basis,
        "years": years,
        "table": table,
    })


@app.route("/api/cities")
@login_required
def api_cities():
    return jsonify(INDIA_CITIES)


@app.route("/api/calculate", methods=["POST"])
@login_required
def api_calculate():
    d = request.get_json(force=True)
    session["last_inputs"] = d

    # ── Step 1: Base Year Emissions ──────────────────────────────────────────
    bldg   = calc_buildings(d)
    trans  = calc_transport(d)
    sw     = calc_solid_waste(d)
    ww     = calc_wastewater(d)
    afolu  = calc_afolu(d)
    ippu   = calc_ippu(d)

    base_by_sector = {
        "Energy Sector": bldg,
        "Transport":     trans,
        "Waste":         sw,
        "Wastewater":    ww,
        "AFOLU":         afolu,
        "IPPU":          ippu,
    }

    base_total = sum(sum(s.values()) for s in base_by_sector.values())

    # ── Step 2: BAU Projections ──────────────────────────────────────────────
    base_year   = int(d.get("base_year",   2025))
    interim1    = int(d.get("interim1",    2030))
    interim2    = int(d.get("interim2",    2040))
    target_year = int(d.get("target_year", 2050))
    years = sorted(set([base_year, interim1, interim2, target_year]))

    bau_by_year = {}
    for yr in years:
        bau_by_year[yr] = calc_bau(base_by_sector, d, yr)

    bau_totals = {yr: sum(sum(s.values()) for s in bau_by_year[yr].values())
                  for yr in years}

    # ── Step 3: Target Setting ───────────────────────────────────────────────
    targets = calc_targets(bau_totals, d)

    # ── Step 4: Scenario Emissions (E&P and High Ambition) ──────────────────
    ep_totals, ha_totals = calc_scenarios(base_by_sector, bau_by_year, d)

    # ── Step 5: Mitigation Budget ────────────────────────────────────────────
    budget_rows, total_inv = calc_mitigation_budget(
        base_by_sector, bau_by_year, ha_totals, d)

    # ── Step 6: Milestones ───────────────────────────────────────────────────
    milestones = calc_milestones(bau_totals, ep_totals, ha_totals, targets, d)

    # ── Step 7: KPIs ─────────────────────────────────────────────────────────
    population = max(float(d.get("population", 1) or 1), 1)
    area_sqkm  = max(float(d.get("area_sqkm",  1) or 1), 1)
    kpis = {
        "base_total_mt": round(base_total / 1e6, 2),
        "per_capita":    round(base_total / population, 2),
        "per_sqkm":      round(base_total / area_sqkm / 1000, 2),
        "bau_end_mt":    round(bau_totals.get(target_year, 0) / 1e6, 2),
        "ep_end_mt":     round(ep_totals.get(target_year, 0) / 1e6, 2),
        "ha_end_mt":     round(ha_totals.get(target_year, 0) / 1e6, 2),
        "target_mt":     round(targets.get(target_year, 0) / 1e6, 2),
        "total_inv":     round(total_inv, 1),
        "base_year":     base_year,
        "target_year":   target_year,
    }

    # ── Step 8: Sector Detail Table ──────────────────────────────────────────
    sector_detail = []
    for sector, subs in base_by_sector.items():
        for sub, val in subs.items():
            if val != 0:
                sector_detail.append({
                    "sector":    f"{sector} – {sub}",
                    "emissions": round(val),
                    "share":     f"{val/base_total*100:.1f}%" if base_total > 0 else "0%"
                })
    sector_detail.sort(key=lambda x: -x["emissions"])

    # ── Step 9: Charts ───────────────────────────────────────────────────────
    charts = {
        "trajectory": make_trajectory_chart(bau_totals, ep_totals, ha_totals, targets, years),
        "pie":        make_pie_chart(base_by_sector),
        "bar_group":  make_bar_chart(bau_totals, ep_totals, ha_totals, targets),
        "budget":     make_budget_chart(budget_rows),
        "subsector":  make_subsector_bar(base_by_sector),
    }

    return jsonify({
        "kpis":          kpis,
        "charts":        charts,
        "milestones":    milestones,
        "sector_detail": sector_detail,
        "budget":        budget_rows,
    })


@app.route("/api/download/excel", methods=["POST"])
@login_required
def download_excel():
    """
    Generate styled Excel export mirroring the structure of the ASCENT workbook.
    4 sheets: Summary, Base Year Emissions, BAU Projections, Mitigation Budget
    """
    d = request.get_json(force=True)

    # Re-run calculations
    bldg  = calc_buildings(d)
    trans = calc_transport(d)
    sw    = calc_solid_waste(d)
    ww    = calc_wastewater(d)
    afolu = calc_afolu(d)
    ippu  = calc_ippu(d)

    base_by_sector = {
        "Energy Sector": bldg, "Transport": trans,
        "Waste": sw, "Wastewater": ww, "AFOLU": afolu, "IPPU": ippu,
    }
    base_total = sum(sum(s.values()) for s in base_by_sector.values())
    population  = max(float(d.get("population", 1) or 1), 1)
    area_sqkm   = max(float(d.get("area_sqkm", 1) or 1), 1)

    base_year   = int(d.get("base_year",   2025))
    interim1    = int(d.get("interim1",    2030))
    interim2    = int(d.get("interim2",    2040))
    target_year = int(d.get("target_year", 2050))
    years = sorted(set([base_year, interim1, interim2, target_year]))

    bau_by_year = {yr: calc_bau(base_by_sector, d, yr) for yr in years}
    bau_totals  = {yr: sum(sum(s.values()) for s in bau_by_year[yr].values()) for yr in years}
    targets     = calc_targets(bau_totals, d)
    ep_totals, ha_totals = calc_scenarios(base_by_sector, bau_by_year, d)
    budget_rows, total_inv = calc_mitigation_budget(base_by_sector, bau_by_year, ha_totals, d)

    wb = Workbook()

    HDR_FILL   = PatternFill("solid", fgColor="1a2744")
    TEAL_FILL  = PatternFill("solid", fgColor="00b4a6")
    LIGHT_FILL = PatternFill("solid", fgColor="e8f4f3")
    HDR_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    BOLD       = Font(name="Calibri", bold=True, size=10)
    NORMAL     = Font(name="Calibri", size=10)
    CENTER     = Alignment(horizontal="center", vertical="center")
    LEFT       = Alignment(horizontal="left",   vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    thin_border= Border(left=thin, right=thin, top=thin, bottom=thin)

    city  = d.get("city", d.get("district","City"))
    state = d.get("state","")
    tier  = d.get("tier","District")

    def hdr_row(ws, row, cols, values):
        for c, v in zip(cols, values):
            cell = ws.cell(row=row, column=c, value=v)
            cell.fill = HDR_FILL; cell.font = HDR_FONT
            cell.alignment = CENTER; cell.border = thin_border

    def data_row(ws, row, cols, values, alt=False):
        fill = LIGHT_FILL if alt else PatternFill()
        for c, v in zip(cols, values):
            cell = ws.cell(row=row, column=c, value=v)
            if alt: cell.fill = fill
            cell.font = NORMAL; cell.alignment = LEFT; cell.border = thin_border

    # ── Sheet 1: Summary ───────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 22

    ws1.merge_cells("A1:B1")
    c = ws1["A1"]
    c.value = f"ASCENT GHG Inventory — {city}, {state}"
    c.font = Font(name="Calibri", bold=True, size=14, color="1a2744")
    c.alignment = CENTER
    ws1.row_dimensions[1].height = 28

    meta = [
        ("City / District", city), ("State", state), ("Government Tier", tier),
        ("Population", f"{int(population):,}"),
        ("Area (km²)", area_sqkm),
        ("Base Year", base_year), ("Target Year", target_year),
        ("Climate Zone", d.get("climate", "—")),
        ("Growth Rate (BAU)", f"{float(d.get('growth_rate',0.03))*100:.1f}%"),
        ("Population Growth Rate", f"{float(d.get('pop_growth_rate', 0) or 0):.2f}%"),
        ("Avg. Annual Rainfall (mm/yr)", d.get('annual_rainfall', '—')),
        ("Avg. Min Temperature (°C)", d.get('temp_min', '—')),
        ("Avg. Max Temperature (°C)", d.get('temp_max', '—')),
        ("Avg. Temperature (°C)", d.get('temp_avg', '—')),
        ("GDP (Crore ₹)", d.get('gdp', '—')),
        ("GDP Growth Rate (%)", d.get('gdp_growth_rate', '—')),
        ("Target Reduction", f"{float(d.get('target_pct',0.45))*100:.0f}%"),
        ("", ""),
        ("Base Year Total Emissions (tCO₂e)", round(base_total)),
        ("Per Capita (tCO₂e/person)", round(base_total/population, 2)),
        ("Per km² (tCO₂e/km²)", round(base_total/area_sqkm)),
        ("BAU at Target Year (Mt)", round(bau_totals.get(target_year,0)/1e6, 2)),
        ("E&P at Target Year (Mt)", round(ep_totals.get(target_year,0)/1e6, 2)),
        ("High Ambition at Target Year (Mt)", round(ha_totals.get(target_year,0)/1e6, 2)),
        ("GHG Reduction Needed (Mt)", round((bau_totals.get(target_year,0)-targets.get(target_year,0))/1e6, 2)),
        ("Total Mitigation Investment (₹ Cr)", round(total_inv, 1)),
    ]
    for r, (label, val) in enumerate(meta, start=3):
        ws1.cell(row=r, column=1, value=label).font = BOLD
        ws1.cell(row=r, column=2, value=val).font = NORMAL

    # ── Sheet 2: Base Year Emissions ──────────────────────────────────────
    ws2 = wb.create_sheet("Base Year Emissions")
    for col, w in zip("ABCD", [30, 28, 20, 15]):
        ws2.column_dimensions[chr(ord("A")+["ABCD".index(c) for c in "ABCD"][["ABCD".index(c) for c in "ABCD"].index(col.replace(col,col))])].width = w
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 28
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 14

    hdr_row(ws2, 1, [1,2,3,4],
            ["Sector","Sub-Sector","Emissions (tCO₂e)","Share (%)"])
    r = 2
    for sector, subs in base_by_sector.items():
        for sub, val in subs.items():
            share = val/base_total*100 if base_total > 0 else 0
            data_row(ws2, r, [1,2,3,4],
                     [sector, sub, round(val), f"{share:.1f}%"], alt=(r%2==0))
            r += 1
    data_row(ws2, r, [1,2,3,4],
             ["TOTAL", "", round(base_total), "100%"], alt=False)
    ws2.cell(row=r, column=1).font = BOLD
    ws2.cell(row=r, column=3).font = BOLD

    # ── Sheet 3: BAU & Scenarios ──────────────────────────────────────────
    ws3 = wb.create_sheet("BAU & Scenarios")
    ws3.column_dimensions["A"].width = 10
    for col in "BCDE": ws3.column_dimensions[col].width = 20

    hdr_row(ws3, 1, [1,2,3,4,5],
            ["Year","BAU (Mt CO₂e)","E&P (Mt CO₂e)","High Ambition (Mt)","Target (Mt)"])
    for r, yr in enumerate(years, start=2):
        data_row(ws3, r, [1,2,3,4,5], [
            yr,
            round(bau_totals[yr]/1e6, 3),
            round(ep_totals.get(yr,0)/1e6, 3),
            round(ha_totals.get(yr,0)/1e6, 3),
            round(targets.get(yr,0)/1e6, 3),
        ], alt=(r%2==0))

    # ── Sheet 4: Mitigation Budget ────────────────────────────────────────
    ws4 = wb.create_sheet("Mitigation Budget")
    ws4.column_dimensions["A"].width = 22
    ws4.column_dimensions["B"].width = 20
    ws4.column_dimensions["C"].width = 14
    ws4.column_dimensions["D"].width = 22
    ws4.column_dimensions["E"].width = 22

    hdr_row(ws4, 1, [1,2,3,4,5],
            ["Sector","BAU at Target Year (t)","Reduction %",
             "GHG Reduced (t CO₂e)","Investment (₹ Crore)"])
    for r, row in enumerate(budget_rows, start=2):
        data_row(ws4, r, [1,2,3,4,5], [
            row["Sector"],
            row["BAU (t CO2e)"],
            row["Reduction %"],
            row["GHG Reduced (t CO2e)"],
            row["Investment (Crore)"],
        ], alt=(r%2==0))
        if row["Sector"] == "TOTAL":
            for c in range(1,6):
                ws4.cell(row=r, column=c).font = BOLD
                ws4.cell(row=r, column=c).fill = TEAL_FILL

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"ASCENT_{city.replace(' ','_')}_{base_year}_{target_year}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/download/csv", methods=["POST"])
@login_required
def download_csv():
    d = request.get_json(force=True)
    bldg  = calc_buildings(d); trans = calc_transport(d)
    sw    = calc_solid_waste(d); ww  = calc_wastewater(d)
    afolu = calc_afolu(d); ippu = calc_ippu(d)
    base_by_sector = {
        "Energy Sector": bldg, "Transport": trans,
        "Waste": sw, "Wastewater": ww, "AFOLU": afolu, "IPPU": ippu,
    }
    base_total = sum(sum(s.values()) for s in base_by_sector.values())
    base_year   = int(d.get("base_year",2025))
    interim1    = int(d.get("interim1",2030))
    interim2    = int(d.get("interim2",2040))
    target_year = int(d.get("target_year",2050))
    years = sorted(set([base_year, interim1, interim2, target_year]))
    bau_by_year = {yr: calc_bau(base_by_sector, d, yr) for yr in years}
    bau_totals  = {yr: sum(sum(s.values()) for s in bau_by_year[yr].values()) for yr in years}
    targets     = calc_targets(bau_totals, d)
    ep_totals, ha_totals = calc_scenarios(base_by_sector, bau_by_year, d)

    lines = ["Year,BAU (Mt),E&P (Mt),High Ambition (Mt),Target (Mt)"]
    for yr in years:
        lines.append(f"{yr},"
                     f"{bau_totals[yr]/1e6:.3f},"
                     f"{ep_totals.get(yr,0)/1e6:.3f},"
                     f"{ha_totals.get(yr,0)/1e6:.3f},"
                     f"{targets.get(yr,0)/1e6:.3f}")
    csv_text = "\n".join(lines)
    buf = io.BytesIO(csv_text.encode())
    buf.seek(0)
    city = d.get("city", d.get("district","city")).replace(" ","_")
    return send_file(buf, as_attachment=True,
                     download_name=f"ASCENT_{city}_scenarios.csv",
                     mimetype="text/csv")

# ═══════════════════════════════════════════════════════════════════════════════
# SUB-PAGE ROUTES  (9 detail pages mirroring Excel dashboard sheets)
# ═══════════════════════════════════════════════════════════════════════════════
 
PAGE_META = [
    ("emission-profile",  "Base-Emission Profile District"),
    ("base-inventory",    "Base Year GHG Inventory"),
    ("bau-scenario",      "BAU Scenario"),
    ("bau-district",      "Dashboard BAU District"),
    ("target-setting",    "Target Setting"),
    ("ep-scenario",       "E&P Scenario"),
    ("ha-scenario",       "High Ambition Scenario"),
    ("emission-graph",    "Emission Reduction Graph"),
    ("scenario-compare",  "Dashboard Scenario Comparison"),
]
 
# Register one route per page using a closure so endpoint names are unique
def _register_page_routes():
    for slug, title in PAGE_META:
        def _make_view(s=slug, t=title):
            @login_required
            def view_fn():
                return render_template("pages/page_base.html",
                                       page_slug=s, page_title=t)
            view_fn.__name__ = f"page_{s.replace('-', '_')}"
            return view_fn
        app.add_url_rule(f"/pages/{slug}",
                         endpoint=f"page_{slug.replace('-','_')}",
                         view_func=_make_view())
 
_register_page_routes()
 
 
# ─── Shared calculation helper ─────────────────────────────────────────────────
def _run_full_calc(d):
    """Run the complete calculation pipeline and return all intermediate results."""
    bldg  = calc_buildings(d)
    trans = calc_transport(d)
    sw    = calc_solid_waste(d)
    ww    = calc_wastewater(d)
    afolu = calc_afolu(d)
    ippu  = calc_ippu(d)
 
    base_by_sector = {
        "Energy Sector": bldg,
        "Transport":     trans,
        "Waste":         sw,
        "Wastewater":    ww,
        "AFOLU":         afolu,
        "IPPU":          ippu,
    }
    base_total  = sum(sum(s.values()) for s in base_by_sector.values())
    base_year   = int(d.get("base_year",   2025))
    interim1    = int(d.get("interim1",    2030))
    interim2    = int(d.get("interim2",    2040))
    target_year = int(d.get("target_year", 2050))
    years       = sorted({base_year, interim1, interim2, target_year})
 
    bau_by_year = {yr: calc_bau(base_by_sector, d, yr) for yr in years}
    bau_totals  = {yr: sum(sum(s.values()) for s in bau_by_year[yr].values())
                   for yr in years}
    targets             = calc_targets(bau_totals, d)
    ep_totals, ha_totals = calc_scenarios(base_by_sector, bau_by_year, d)
    budget_rows, total_inv = calc_mitigation_budget(
        base_by_sector, bau_by_year, ha_totals, d)
    milestones = calc_milestones(bau_totals, ep_totals, ha_totals, targets, d)
 
    return dict(
        base_by_sector=base_by_sector,
        base_total=base_total,
        base_year=base_year,
        target_year=target_year,
        years=years,
        bau_by_year=bau_by_year,
        bau_totals=bau_totals,
        targets=targets,
        ep_totals=ep_totals,
        ha_totals=ha_totals,
        budget_rows=budget_rows,
        total_inv=total_inv,
        milestones=milestones,
        d=d,
    )
 
 
# ─── /api/page-data/<slug>  POST ──────────────────────────────────────────────
@app.route("/api/page-data/<page_slug>", methods=["POST"])
@login_required
def api_page_data(page_slug):
    """Single endpoint that returns JSON for whichever sub-page requests it."""
    try:
        d = request.get_json(force=True) or session.get("last_inputs")
        if not d:
            return jsonify({"error": "No input data found. Please go back to the Results page and ensure a calculation has been run first."}), 400
 
        c = _run_full_calc(d)   # c = calc results dict
 
        dispatch = {
            "emission-profile": _page_emission_profile,
            "base-inventory":   _page_base_inventory,
            "bau-scenario":     _page_bau_scenario,
            "bau-district":     _page_bau_district,
            "target-setting":   _page_target_setting,
            "ep-scenario":      _page_ep_scenario,
            "ha-scenario":      _page_ha_scenario,
            "emission-graph":   _page_emission_graph,
            "scenario-compare": _page_scenario_compare,
        }
        fn = dispatch.get(page_slug)
        if fn is None:
            return jsonify({"error": f"Unknown page: {page_slug}"}), 404
 
        return jsonify(fn(c))
 
    except Exception as ex:
        import traceback
        return jsonify({"error": str(ex), "trace": traceback.format_exc()}), 500
 
 
# ─── Page data builders ────────────────────────────────────────────────────────
 
def _page_emission_profile(c):
    base_by_sector = c["base_by_sector"]
    base_total     = c["base_total"]
    d              = c["d"]
    population     = max(float(d.get("population", 1) or 1), 1)
    area           = max(float(d.get("area_sqkm",  1) or 1), 1)
 
    rows = []
    for sector, subs in base_by_sector.items():
        for sub, val in subs.items():
            rows.append({
                "sector":           sector,
                "subsector":        sub,
                "emissions_tco2e":  round(val),
                "share_pct":        round(val / base_total * 100, 2) if base_total else 0,
                "per_capita":       round(val / population, 4),
                "per_sqkm":         round(val / area, 2),
            })
    rows.sort(key=lambda x: -x["emissions_tco2e"])
 
    return {
        "profile_rows": rows,
        "total":        round(base_total),
        "per_capita":   round(base_total / population, 2),
        "per_sqkm":     round(base_total / area, 2),
        "base_year":    c["base_year"],
        "city":         d.get("city", d.get("district", "City")),
        "state":        d.get("state", ""),
        "chart_pie":    make_pie_chart(base_by_sector),
        "chart_bar":    make_subsector_bar(base_by_sector),
    }
 
 
def _page_base_inventory(c):
    base_by_sector = c["base_by_sector"]
    base_total     = c["base_total"]
    rows = []
    for sector, subs in base_by_sector.items():
        sector_total = sum(subs.values())
        for sub, val in subs.items():
            rows.append({
                "sector":        sector,
                "subsector":     sub,
                "co2_eq":        round(val),
                "sector_share":  f"{val/sector_total*100:.1f}%" if sector_total else "0%",
                "total_share":   f"{val/base_total*100:.1f}%" if base_total else "0%",
            })
        # subtotal row
        rows.append({
            "sector":        sector,
            "subsector":     "── SUBTOTAL",
            "co2_eq":        round(sector_total),
            "sector_share":  "100%",
            "total_share":   f"{sector_total/base_total*100:.1f}%" if base_total else "0%",
            "is_subtotal":   True,
        })
    return {
        "inventory_rows": rows,
        "total":          round(base_total),
        "base_year":      c["base_year"],
        "city":           c["d"].get("city", c["d"].get("district", "City")),
        "state":          c["d"].get("state", ""),
        "chart_bar":      make_subsector_bar(base_by_sector),
    }
 
 
def _page_bau_scenario(c):
    bau_by_year = c["bau_by_year"]
    bau_totals  = c["bau_totals"]
    years       = c["years"]
    rows = []
    for yr in years:
        row = {"year": yr, "total_mt": round(bau_totals[yr] / 1e6, 3)}
        for sector, subs in bau_by_year[yr].items():
            row[sector] = round(sum(subs.values()) / 1e6, 3)
        rows.append(row)
    sectors = list(c["base_by_sector"].keys())
    return {
        "bau_rows":   rows,
        "sectors":    sectors,
        "years":      years,
        "city":       c["d"].get("city", c["d"].get("district", "City")),
        "state":      c["d"].get("state", ""),
        "base_year":  c["base_year"],
    }
 
 
def _page_bau_district(c):
    bau_totals  = c["bau_totals"]
    bau_by_year = c["bau_by_year"]
    years       = c["years"]
    base_year   = c["base_year"]
    population  = max(float(c["d"].get("population", 1) or 1), 1)
    base_total  = bau_totals.get(base_year, 0)
    return {
        "bau_totals":        {str(y): round(v / 1e6, 3) for y, v in bau_totals.items()},
        "per_capita_base":   round(base_total / population, 2),
        "city":              c["d"].get("city", c["d"].get("district", "City")),
        "state":             c["d"].get("state", ""),
        "base_year":         base_year,
        "target_year":       c["target_year"],
        "trajectory_chart":  make_trajectory_chart(
            bau_totals, bau_totals, bau_totals, bau_totals, years),
        "pie_chart":         make_pie_chart(bau_by_year.get(base_year, {})),
    }
 
 
def _page_target_setting(c):
    return {
        "milestones":  c["milestones"],
        "bau":         {str(y): round(v / 1e6, 3) for y, v in c["bau_totals"].items()},
        "targets":     {str(y): round(v / 1e6, 3) for y, v in c["targets"].items()},
        "ep":          {str(y): round(v / 1e6, 3) for y, v in c["ep_totals"].items()},
        "ha":          {str(y): round(v / 1e6, 3) for y, v in c["ha_totals"].items()},
        "city":        c["d"].get("city", c["d"].get("district", "City")),
        "state":       c["d"].get("state", ""),
        "base_year":   c["base_year"],
        "target_year": c["target_year"],
        "target_pct":  float(c["d"].get("target_pct", 0.45) or 0.45) * 100,
    }
 
 
def _page_ep_scenario(c):
    years = c["years"]
    return {
        "ep":    {str(y): round(c["ep_totals"].get(y, 0) / 1e6, 3) for y in years},
        "bau":   {str(y): round(c["bau_totals"].get(y, 0) / 1e6, 3) for y in years},
        "ha":    {str(y): round(c["ha_totals"].get(y, 0) / 1e6, 3) for y in years},
        "city":  c["d"].get("city", c["d"].get("district", "City")),
        "state": c["d"].get("state", ""),
        "years": years,
        "chart": make_trajectory_chart(
            c["bau_totals"], c["ep_totals"],
            c["ha_totals"],  c["targets"], years),
    }
 
 
def _page_ha_scenario(c):
    years = c["years"]
    return {
        "ha":    {str(y): round(c["ha_totals"].get(y, 0) / 1e6, 3) for y in years},
        "bau":   {str(y): round(c["bau_totals"].get(y, 0) / 1e6, 3) for y in years},
        "ep":    {str(y): round(c["ep_totals"].get(y, 0) / 1e6, 3) for y in years},
        "city":  c["d"].get("city", c["d"].get("district", "City")),
        "state": c["d"].get("state", ""),
        "years": years,
        "total_inv": round(c["total_inv"], 1),
        "chart": make_trajectory_chart(
            c["bau_totals"], c["ep_totals"],
            c["ha_totals"],  c["targets"], years),
        "budget_chart": make_budget_chart(c["budget_rows"]),
    }
 
 
def _page_emission_graph(c):
    years = c["years"]
    return {
        "years":    years,
        "bau":      {str(y): round(c["bau_totals"].get(y, 0) / 1e6, 3) for y in years},
        "ep":       {str(y): round(c["ep_totals"].get(y, 0) / 1e6, 3) for y in years},
        "ha":       {str(y): round(c["ha_totals"].get(y, 0) / 1e6, 3) for y in years},
        "target":   {str(y): round(c["targets"].get(y, 0) / 1e6, 3) for y in years},
        "city":     c["d"].get("city", c["d"].get("district", "City")),
        "state":    c["d"].get("state", ""),
        "chart":    make_trajectory_chart(
            c["bau_totals"], c["ep_totals"],
            c["ha_totals"],  c["targets"], years),
    }
 
 
def _page_scenario_compare(c):
    years = c["years"]
    return {
        "years":      years,
        "bau":        {str(y): round(c["bau_totals"].get(y, 0) / 1e6, 3) for y in years},
        "ep":         {str(y): round(c["ep_totals"].get(y, 0) / 1e6, 3) for y in years},
        "ha":         {str(y): round(c["ha_totals"].get(y, 0) / 1e6, 3) for y in years},
        "target":     {str(y): round(c["targets"].get(y, 0) / 1e6, 3) for y in years},
        "city":       c["d"].get("city", c["d"].get("district", "City")),
        "state":      c["d"].get("state", ""),
        "total_inv":  round(c["total_inv"], 1),
        "budget":     c["budget_rows"],
        "bar_chart":  make_bar_chart(
            c["bau_totals"], c["ep_totals"],
            c["ha_totals"],  c["targets"]),
        "trajectory": make_trajectory_chart(
            c["bau_totals"], c["ep_totals"],
            c["ha_totals"],  c["targets"], years),
    }

# ═══════════════════════════════════════════════════════════════════════════════
# CONVERSION FACTOR API (for the expanded questionnaire)
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/conv-factors")
@login_required
def api_conv_factors():
    """Return the fuel × unit → TJ table as JSON so the front-end can show
    valid unit options for each fuel, and optionally preview TJ conversions."""
    # Group by fuel name; values become list of allowed units
    by_fuel = {}
    for (fuel, unit), factor in CONV_FACTORS.items():
        by_fuel.setdefault(fuel, {})[unit] = factor
    return jsonify({"factors": by_fuel, "fuels": sorted(by_fuel.keys())})


# ═══════════════════════════════════════════════════════════════════════════════
# AUTHENTICATION ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/login", methods=["GET"])
def login_page():
    # If already signed in, send them to the questionnaire.
    if current_user():
        return redirect(url_for("index"))
    return render_template("login.html")


@app.route("/auth/register", methods=["POST"])
def auth_register():
    data     = request.get_json(silent=True) or {}
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""

    if not _USERNAME_RE.match(username):
        return jsonify({
            "ok": False,
            "error": "Username must be 3-32 chars: letters, digits, _ . -"
        }), 400
    if len(password) < 1:
        return jsonify({"ok": False, "error": "Password is required."}), 400

    with _STORE_LOCK:
        users = _load_users()
        if username in users:
            return jsonify({
                "ok": False,
                "error": "That username is already taken. Try signing in."
            }), 409
        users[username] = generate_password_hash(password)
        _save_users(users)

    session["username"] = username
    return jsonify({"ok": True, "redirect": url_for("index")})


@app.route("/auth/login", methods=["POST"])
def auth_login():
    data     = request.get_json(silent=True) or {}
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""

    if not username or not password:
        return jsonify({"ok": False, "error": "Username and password are required."}), 400

    users = _load_users()
    stored = users.get(username)
    if not stored or not check_password_hash(stored, password):
        return jsonify({"ok": False, "error": "Invalid username or password."}), 401

    session["username"] = username
    nxt = request.args.get("next") or url_for("index")
    # Only allow same-origin relative redirects
    if not nxt.startswith("/"):
        nxt = url_for("index")
    return jsonify({"ok": True, "redirect": nxt})


@app.route("/logout", methods=["GET", "POST"])
def logout():
    session.pop("username", None)
    return redirect(url_for("login_page"))


# ═══════════════════════════════════════════════════════════════════════════════
# PER-USER FORM DATA  (browsers + sessions share state via the account)
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/user-data", methods=["GET"])
@login_required
def user_data_get():
    user = current_user()
    return jsonify(_load_user_form(user) or {"form": {}})


@app.route("/api/user-data", methods=["POST"])
@login_required
def user_data_save():
    user    = current_user()
    payload = request.get_json(silent=True) or {}
    # Only persist the "form" sub-object — never trust arbitrary top-level keys.
    form    = payload.get("form")
    if not isinstance(form, dict):
        return jsonify({"ok": False, "error": "Expected {\"form\": {...}}"}), 400
    _save_user_form(user, {"form": form})
    return jsonify({"ok": True})


@app.route("/api/user-data", methods=["DELETE"])
@login_required
def user_data_delete():
    user = current_user()
    _delete_user_form(user)
    return jsonify({"ok": True})


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
